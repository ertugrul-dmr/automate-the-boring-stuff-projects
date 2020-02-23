'''
A program blankRowInserter.py that takes two integers and a filename string as command line arguments.
First integer N and the second integer M.
Starting at row N, the program should insert M blank rows into the spreadsheet.
'''
import openpyxl

inputWB = openpyxl.load_workbook('produceSales.xlsx', data_only=True)
inputSheet = inputWB.active

maxRow = inputSheet.max_row
maxColumn = inputSheet.max_column

outputwb = openpyxl.Workbook()
outputSheet = outputwb.active

N = 3
M = 4


for r in range(1, maxRow + 1):
    for c in range(1, maxColumn + 1):
        if r <= N:
            outputSheet.cell(row=r, column=c).value = inputSheet.cell(
                row=r, column=c).value
        if r > N:
            outputSheet.cell(row=r + M, column=c).value = inputSheet.cell(
                row=r, column=c).value
outputwb.save('blanked.xlsx')
