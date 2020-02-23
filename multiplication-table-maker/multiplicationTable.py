#! python3
# A program that takes a number N from the command line and creates an N×N multiplication table in an Excel spreadsheet.
import openpyxl
import sys
import os
from openpyxl.styles import Font

x = int(sys.argv[1])
wb = openpyxl.Workbook()
sheet = wb.active
bold = Font(bold=True)

for c in range(1, x + 1):
    sheet.cell(column=1, row=c + 1).font = bold
    sheet.cell(column=c + 1, row=1).font = bold
    sheet.cell(column=1, row=c + 1).value = c
    sheet.cell(column=c + 1, row=1).value = c
    for r in range(1, x + 1):
        sheet.cell(column=c + 1, row=r + 1).value = (r) * (c)

print('Calculating')

wb.save('multiplicationtable.xlsx')
