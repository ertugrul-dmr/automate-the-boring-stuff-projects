# A program that reads all the Excel files in the current working directory and outputs them as CSV files.
import csv
import openpyxl
import os
for excelFile in os.listdir('your xlsx files folder path'):
    # Skip non-xlsx files, load the workbook object.
    if excelFile.endswith('.xlsx'):
        wb = openpyxl.load_workbook(excelFile)
        for sheetName in wb.get_sheet_names():
            # Loop through every sheet in the workbook.
            sheet = wb.get_sheet_by_name(sheetName)

            # Create the CSV filename from the Excel filename and sheet title.
            csvFile = open(f'{excelFile}_{sheetName}.csv', 'w', newline='')
            # Create the csv.writer object for this CSV file.
            outputWriter = csv.writer(csvFile)

            # Loop through every row in the sheet.
            for rowNum in range(1, sheet.max_row + 1):
                rowData = []    # append each cell to this list
                # Loop through each cell in the row.
                for colNum in range(1, sheet.max_column + 1):
                        # Append each cell's data to rowData.
                    x = sheet.cell(row=rowNum, column=colNum).value
                    rowData.append(x)

                # Write the rowData list to the CSV file.
                outputWriter.writerow(rowData)

            csvFile.close()
