# A program to invert the row and column of the cells in the spreadsheet.

import openpyxl

inputwb = openpyxl.load_workbook('produceSales.xlsx', data_only=True)
inputSheet = inputwb.active

maxRow = inputSheet.max_row
maxColumn = inputSheet.max_column

outputwb = openpyxl.Workbook()
outputSheet = outputwb.active

for r in range(1, maxRow + 1):
    for c in range(1, maxColumn + 1):
        outputSheet.cell(row=c, column=r).value = inputSheet.cell(
            row=r, column=c).value

outputwb.save('inverted.xlsx')
