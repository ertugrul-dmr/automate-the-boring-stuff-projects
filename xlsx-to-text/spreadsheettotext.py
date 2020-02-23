# The program should open a spreadsheet and write the cells of column A into one text file, the cells of column B into another text file, and so on.
import openpyxl

wb = openpyxl.load_workbook('spread.xlsx')
sheet = wb.active

maxRow = sheet.max_row

for i in range(maxRow):
    with open(f'output {i+1}.txt', 'w') as file:
        file.write(sheet.cell(row=i + 1, column=1).value)
